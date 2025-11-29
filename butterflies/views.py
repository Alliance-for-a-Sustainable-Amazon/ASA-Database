# views.py

# --- Imports ---
from xml.parsers.expat import model
from django.apps import apps
from django.shortcuts import render, redirect
from django.views.generic import ListView
from django.http import Http404, HttpResponse, HttpResponseRedirect
from django.forms import modelform_factory
from django.contrib import messages
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.db import models
from django.template.loader import get_template

import csv
import io
import openpyxl
import os
import pandas as pd
import re
import datetime
import tempfile
from dateutil import parser as date_parser
from operator import itemgetter
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import json
import pickle
import base64
import logging

# App-specific imports
from .models import Specimen, Locality, Initials
from .forms import SpecimenForm, SpecimenEditForm, LocalityForm, InitialsForm
from .utils import dot_if_none
from butterflies.utils.image_utils import get_specimen_image_urls
from .auth_utils import admin_required, guest_allowed, is_guest_mode
from .filter_utils import FilterBuilder, apply_model_filters
from .utils.image_utils import get_specimen_image_urls_batch


# --- Utility Functions ---
def model_list():
    """
    Return all models in the butterflies app.
    """
    return list(apps.get_app_config('butterflies').get_models())

def parse_date_value(value):
    """
    Parse a value into a date object.
    Complement to format_date_value()
    
    Parameters:
        value: Date value in any format (string, date, datetime)
    Returns:
        date object or None if parsing fails
    """
    if not value:
        return None
        
    try:
        # If already a datetime, convert to date
        if isinstance(value, datetime.datetime):
            return value.date()
        
        # If it's already a date object, return it
        elif isinstance(value, datetime.date):
            return value
            
        # Try direct parsing of standard format first (for string values from user input)
        standard_pattern = r"^\d{1,2} [A-Za-z]+,? \d{4}$"
        if isinstance(value, str) and re.match(standard_pattern, value):
            # Try a few variations of our standard format
            for fmt in ["%d %B, %Y", "%d %B %Y"]:
                try:
                    return datetime.datetime.strptime(value, fmt).date()
                except:
                    continue
                    
        # Use dateutil's flexible parser as a fallback for strings
        if isinstance(value, str):
            return date_parser.parse(value).date()
        
        # If we get here with a non-string, non-date value, try to convert
        return datetime.datetime.fromisoformat(str(value)).date()
            
    except Exception as e:
        logging.warning(f"Date parsing error for '{value}' ({type(value)}): {str(e)}")
        return None

def format_date_value(value):
    """
    Format a date value to the standardized string format: DD Month, YYYY.
    Handles datetime objects, date objects, strings, and pandas Timestamp objects.
    
    This is the central date formatting function for the entire application.
    All dates should pass through this function for standardization when displaying.
    
    Parameters:
        value: A date value in any format (datetime, date, string, Timestamp, etc.)
    Returns:
        String in format "DD Month, YYYY" or None if conversion fails
    """
    if pd.isna(value) or value is None or value == '':
        return None
    
    try:
        # Function to format date consistently across platforms
        def format_date(dt):
            # Use %d and strip leading zero instead of %-d which doesn't work on Windows
            day = dt.strftime("%d").lstrip("0")
            if day == "": day = "1"  # Handle case where day is "01"
            month = dt.strftime("%B")  # Full month name
            year = dt.strftime("%Y")   # 4-digit year
            return f"{day} {month}, {year}"
        
        # If value is already a datetime or date object
        if isinstance(value, (datetime.datetime, datetime.date)):
            return format_date(value)
        
        # If value is a pandas Timestamp
        elif hasattr(value, 'to_pydatetime'):
            return format_date(value.to_pydatetime())
            
        # If value is a string, try to parse it
        elif isinstance(value, str):
            # First check if it's already in the correct format (DD Month, YYYY)
            date_pattern = r"^\d{1,2} [A-Za-z]+\.?,? \d{4}$"
            if re.match(date_pattern, value):
                # It's already in the right format, but standardize to ensure consistency
                try:
                    dt = date_parser.parse(value)
                    return format_date(dt)
                except:
                    # If parsing fails, return the original as it's close enough
                    return value
                
            # Next check for ISO format (YYYY-MM-DD)
            iso_pattern = r"^\d{4}-\d{2}-\d{2}$"
            if re.match(iso_pattern, value):
                try:
                    dt = datetime.datetime.strptime(value, "%Y-%m-%d")
                    return format_date(dt)
                except:
                    pass
                
            # Check for DD-MMM-YYYY format (from help text)
            dmm_pattern = r"^\d{1,2}-[A-Za-z]{3,9}-\d{4}$"
            if re.match(dmm_pattern, value):
                try:
                    # Replace abbreviated month names with standard ones if needed
                    value = value.replace(".", "")  # Remove periods from abbreviated months
                    # Try various formats with different separators and month formats
                    for fmt in ["%d-%b-%Y", "%d-%B-%Y"]:
                        try:
                            dt = datetime.datetime.strptime(value, fmt)
                            return format_date(dt)
                        except:
                            pass
                except:
                    pass
                    
            # Try some common date formats
            common_formats = [
                "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y",  # Slash formats
                "%Y.%m.%d", "%d.%m.%Y", "%m.%d.%Y",  # Dot formats
                "%b %d %Y", "%B %d %Y", "%d %b %Y", "%d %B %Y",  # Word month formats
                "%Y%m%d"  # No separator format
            ]
            
            for fmt in common_formats:
                try:
                    dt = datetime.datetime.strptime(value, fmt)
                    return format_date(dt)
                except:
                    continue
            
            # Last resort: Try dateutil's flexible parser
            try:
                dt = date_parser.parse(value, dayfirst=False)  # default to MM/DD/YYYY
                return format_date(dt)
            except:
                # Try again with day first
                try:
                    dt = date_parser.parse(value, dayfirst=True)  # Try DD/MM/YYYY
                    return format_date(dt)
                except:
                    pass
                
    except Exception as e:
        logging.warning(f"Date conversion error for value '{value}': {str(e)}")
        
    # If all conversions fail, return the original value if it's a string
    return str(value) if value is not None else None

# --- Generic CRUD Views ---

@guest_allowed
def dynamic_list(request, model_name):
    """
    Generic dynamic list view for any model.
    Displays all objects for the specified model with filtering capabilities.
    Includes pagination for better performance.
    Supports both table and grid view modes for specimens.
    Allows guest access for read-only viewing.
    
    Parameters:
        request: HTTP request object
        model_name: String name of the model to list
    Returns:
        Rendered template with paginated and filtered objects
    """    
    model = None
    for m in model_list():
        if m._meta.model_name == model_name:
            model = m
            break
    if not model:
        raise Http404("Model not found")
    
    # Handle view mode toggle for specimen model
    view_mode = 'table'  # default view mode
    if model_name == 'specimen':
        # Check if view mode is being changed
        if request.method == 'POST' and 'toggle_view_mode' in request.POST:
            new_view_mode = request.POST.get('view_mode', 'table')
            request.session[f'{model_name}_view_mode'] = new_view_mode
            request.session.modified = True
            # Redirect to avoid form resubmission on refresh
            return HttpResponseRedirect(request.get_full_path())
        
        # Get view mode from session or default to table
        view_mode = request.session.get(f'{model_name}_view_mode', 'table')
    
    # Base queryset
    # objects = model.objects.all()
    if model._meta.model_name == 'specimen':
        objects = model.objects.select_related('locality', 'recordedBy', 'georeferencedBy', 'identifiedBy').all()
    else:
        objects = model.objects.all()
    
    # Define special filters for specimen model
    special_filters = None
    if model._meta.model_name == 'specimen':
        special_filters = {
            'catalogNumber': {'field': 'catalogNumber', 'range_support': True},
            'locality': {'field': 'locality__localityCode', 'range_support': False},
            'specimenNumber': {'field': 'specimenNumber', 'range_support': True},
            'year': {'field': 'year', 'range_support': True}
        }
    
    # Apply all filters using the standardized filter utility
    objects = apply_model_filters(objects, model, request, special_filters)
    
    # Fields for display and filtering
    fields = [
        {'name': field.name, 'verbose_name': getattr(field, 'verbose_name', field.name)}
        for field in model._meta.fields
        if field.name != 'id'
    ]
    
    # Add model_name_internal to each object for URL generation
    obj_list = []
    for obj in objects:
        obj.model_name_internal = model._meta.model_name
        obj_list.append(obj)
    
    # Cursor (keyset) pagination - before loading images for efficiency
    class CursorPage:
        def __init__(self, object_list, has_next, has_prev, next_cursor, prev_cursor):
            self.object_list = object_list
            self.has_next = has_next
            self.has_previous = has_prev
            self.next_cursor = next_cursor
            self.prev_cursor = prev_cursor
        # Compatibility with templates that iterate page_obj
        def __iter__(self):
            return iter(self.object_list)
        def __len__(self):
            return len(self.object_list)

    # Choose a stable indexed ordering field for keyset pagination
    # Prefer primary key; for Specimen, catalogNumber is also stable
    if model._meta.model_name == 'specimen':
        order_field = 'catalogNumber'
    else:
        order_field = model._meta.pk.name

    items_per_page = 20
    try:
        current_offset = int(request.GET.get('offset', '0'))
        if current_offset < 0:
            current_offset = 0
    except ValueError:
        current_offset = 0
    after = request.GET.get('after')  # cursor value (last seen key)
    direction = request.GET.get('dir', 'forward')  # 'forward' or 'back'

    # Build base queryset with ordering and apply cursor filter
    if direction == 'back':
        qs = objects.order_by(f'-{order_field}')
        if after:
            try:
                qs = qs.filter(**{f"{order_field}__lt": after})
            except Exception:
                pass
    else:
        qs = objects.order_by(order_field)
        if after:
            try:
                qs = qs.filter(**{f"{order_field}__gt": after})
            except Exception:
                pass

    # Fetch one extra to determine has_next
    page_items = list(qs[:items_per_page + 1])
    has_next = len(page_items) > items_per_page
    if has_next:
        page_items = page_items[:items_per_page]
    # If we fetched in reverse (back direction), flip to display ascending
    if direction == 'back':
        page_items = list(reversed(page_items))

    # Determine prev/next cursors
    next_cursor = None
    prev_cursor = None
    if page_items:
        next_cursor = getattr(page_items[-1], order_field, None)
        prev_cursor = getattr(page_items[0], order_field, None)

    # Ensure objects have model_name_internal for URL generation in templates
    for obj in page_items:
        obj.model_name_internal = model._meta.model_name

    page_obj = CursorPage(page_items, has_next=has_next, has_prev=bool(after), next_cursor=next_cursor, prev_cursor=prev_cursor)

    # Total count for display
    total_count = objects.count()

    # Build clean next/prev URLs using Django QueryDict
    def build_cursor_url(cursor_value, direction_value, new_offset):
        params = request.GET.copy()
        for key in ['after', 'dir', 'page']:
            if key in params:
                del params[key]
        params['after'] = cursor_value
        params['dir'] = direction_value
        params['offset'] = str(new_offset)
        return f"{request.path}?{params.urlencode()}"

    # Compute start/end counters for display
    start_index = current_offset + 1 if page_items else 0
    end_index = current_offset + len(page_items)

    next_offset = current_offset + len(page_items)
    prev_offset = max(current_offset - len(page_items), 0)

    next_url = build_cursor_url(next_cursor, 'forward', next_offset) if (has_next and next_cursor) else None
    prev_url = build_cursor_url(prev_cursor, 'back', prev_offset) if (bool(after) and prev_cursor) else None
    
    # Only load images for the current page objects
    if model_name == 'specimen' and view_mode == 'grid':
        # Get catalog numbers for batch processing
        catalog_numbers = [obj.catalogNumber for obj in page_obj if obj.catalogNumber]
        
        # Use batch loading for grid view
        batch_image_results = get_specimen_image_urls_batch(catalog_numbers)
        
        for obj in page_obj:
            obj.dorsal_image_url = None
            obj.ventral_image_url = None
            if obj.catalogNumber in batch_image_results:
                image_dict = batch_image_results[obj.catalogNumber]
                
                # Get both dorsal and ventral images
                if image_dict.get('dorsal') and image_dict['dorsal'] != 'no data':
                    obj.dorsal_image_url = image_dict['dorsal']
                if image_dict.get('ventral') and image_dict['ventral'] != 'no data':
                    obj.ventral_image_url = image_dict['ventral']
    
    # Check for import errors in session
    import_errors = None
    import_complete = request.session.get('import_complete', False)
    
    if import_complete:
        import_errors = request.session.get('import_errors', [])
        # Add a summary message if there are errors
        if import_errors and len(import_errors) > 0:
            messages.warning(request, f'There were {len(import_errors)} errors during import. See details below.')
        
        # Clear the session variables after retrieving them
        request.session['import_complete'] = False
        if 'import_errors' in request.session:
            del request.session['import_errors']
        
        # Make sure changes are saved to session
        request.session.modified = True
    
    # Create a list of field names for use in the template
    field_names = [field['name'] for field in fields]
    
    return render(request, 'butterflies/dynamic_list.html', {
        'objects': page_obj,
        'fields': fields,
        'field_names': field_names,
        'model_name': model._meta.verbose_name.title(),
        'model_name_internal': model._meta.model_name,
        'request': request,
        'home_url': 'report_table',
        'import_errors': import_errors,  # Pass import errors to the template
        # Provide cursor pagination metadata for templates
        'page_obj': page_obj,
        'cursor_order_field': order_field,
        'cursor_next': getattr(page_obj, 'next_cursor', None),
        'cursor_prev': getattr(page_obj, 'prev_cursor', None),
        'view_mode': view_mode,  # Pass current view mode to template
        'supports_grid_view': model_name == 'specimen',  # Only specimen supports grid view
        'total_count': total_count,
        'next_url': next_url,
        'prev_url': prev_url,
        'start_index': start_index,
        'end_index': end_index,
    })

@guest_allowed
def dynamic_detail(request, model_name, object_id):
    """
    Generic detail view for any model.
    Shows detailed information for a specific object.
    Allows guest access for read-only viewing.
    
    Parameters:
        request: HTTP request object
        model_name: String name of the model
        object_id: Primary key of the object to display
    Returns:
        Rendered template with object details
    """
    model = None
    for m in model_list():
        if m._meta.model_name == model_name:
            model = m
            break
    if not model:
        raise Http404("Model not found")
    try:
        obj = model.objects.get(pk=object_id)
    except model.DoesNotExist:
        raise Http404("Object not found")
    fields = [
        {'name': field.name, 'verbose_name': getattr(field, 'verbose_name', field.name)}
        for field in model._meta.fields
        if field.name != 'id'
    ]
    obj.model_name_internal = model._meta.model_name
    # If specimen, add image URLs
    image_urls = None
    if model._meta.model_name == 'specimen':
        image_urls = get_specimen_image_urls(obj.catalogNumber)
    return render(request, 'butterflies/_detail.html', {
        'object': obj,
        'fields': fields,
        'model_name': model._meta.verbose_name.title(),
        'image_urls': image_urls,
    })

@admin_required
def dynamic_delete(request, model_name, object_id):
    """
    Generic delete view for any model.
    Shows confirmation page before deleting an object.
    Requires admin privileges.
    
    Parameters:
        request: HTTP request object
        model_name: String name of the model
        object_id: Primary key of the object to delete
    Returns:
        Redirect to report_table on successful deletion
        or confirmation template
    """
    model = None
    for m in model_list():
        if m._meta.model_name == model_name:
            model = m
            break
    if not model:
        raise Http404("Model not found")
    try:
        obj = model.objects.get(pk=object_id)
    except model.DoesNotExist:
        raise Http404("Object not found")
    if request.method == 'POST':
        obj.delete()
        from django.urls import reverse
        return redirect(reverse('report_table'))
    obj.model_name_internal = model._meta.model_name
    return render(request, 'butterflies/_detail.html', {
        'object': obj,
        'fields': [
            {'name': field.name, 'verbose_name': getattr(field, 'verbose_name', field.name)}
            for field in model._meta.fields
            if field.name != 'id'
        ],
        'model_name': model._meta.verbose_name.title(),
        'delete_confirm': True,
    })

@login_required
def dynamic_create_edit(request, model_name, object_id=None):
    """
    Generic create/edit view for any model using the organized form approach.
    Uses the organized form template for Specimen model.
    Requires user authentication.
    
    Parameters:
        request: HTTP request object
        model_name: String name of the model to create or edit
        object_id: Optional primary key for edit mode
    Returns:
        Rendered form template or redirect to report_table after save
    """
    model = None
    for m in model_list():
        if m._meta.model_name == model_name:
            model = m
            break
    if not model:
        raise Http404("Model not found")
    
    # Check if restricted model (locality or initials) and user has admin permissions
    if model._meta.model_name in ['locality', 'initials']:
        # Check if user is admin or superuser
        if not (request.user.is_superuser or request.user.groups.filter(name='Admin').exists()):
            return render(request, 'butterflies/auth/access_denied.html')
    
    # Use appropriate form class based on model
    if model._meta.model_name == 'specimen':
        # Use SpecimenEditForm for editing, SpecimenForm for creating
        if object_id:
            form_class = SpecimenEditForm
        else:
            form_class = SpecimenForm
    elif model._meta.model_name == 'locality':
        form_class = LocalityForm
    elif model._meta.model_name == 'initials':
        form_class = InitialsForm
    else:
        form_class = modelform_factory(model, exclude=['id'])
    
    instance = None
    if object_id:
        try:
            instance = model.objects.get(pk=object_id)
        except model.DoesNotExist:
            raise Http404("Object not found")
    
    if request.method == 'POST':
        form = form_class(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            # Redirect to home after save
            return redirect(reverse('report_table'))
    else:
        form = form_class(instance=instance)
    
    # Use model-specific organized template if it's a specimen
    if model._meta.model_name == 'specimen':
        template_name = 'butterflies/specimen_form.html'
    else:
        # Use model-specific template if it exists
        template_name = f'butterflies/{model._meta.model_name}_form.html'
        try:
            get_template(template_name)
        except:
            # Fall back to generic form template
            template_name = 'butterflies/_form.html'
    
    return render(request, template_name, {
        'form': form, 
        'model_name': model._meta.verbose_name.title()
    })

# --- Model-Specific Create Views ---

@login_required
def create_specimen(request):
    """
    Create a new specimen record.
    Uses a dedicated form for specimen creation.
    Requires user authentication.
    
    Parameters:
        request: HTTP request object
    Returns:
        Rendered form template or redirect on success
    """
    form = SpecimenForm()
    if request.method == 'POST':
        form = SpecimenForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Specimen added successfully!')
            form = SpecimenForm()
    return render(request, 'butterflies/specimen_form.html', {'form': form})

@admin_required
def create_locality(request):
    """
    Create a new locality record.
    Uses a form based on the Locality model.
    Requires admin privileges.
    
    Parameters:
        request: HTTP request object
    Returns:
        Rendered form template or redirect on success
    """
    form = LocalityForm()
    if request.method == 'POST':
        form = LocalityForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Locality added successfully!')
            form = LocalityForm()
    return render(request, 'butterflies/locality_form.html', {'form': form})

@admin_required
def create_initials(request):
    """
    Create a new Initials object.
    Handles both form display and processing.
    Requires admin privileges.
    
    Parameters:
        request: HTTP request object
    Returns:
        Rendered form template with success message on submission
    """
    form = InitialsForm()
    if request.method == 'POST':
        form = InitialsForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Initials added successfully!')
            form = InitialsForm()
    return render(request, 'butterflies/initials_form.html', {'form': form})



# --- Report Views ---

@guest_allowed
def report_table(request):
    """
    Shows a paginated table of specimens based on eventDate and eventTime.
    This is the main dashboard view with pagination for better performance.
    Optimized to use database-level sorting instead of Python sorting.
    Requires user authentication.
    
    Parameters:
        request: HTTP request object
    Returns:
        Rendered template with paginated specimens
    """
    from django.db.models import F
    
    # Get top 20 most recent specimens using database-level ordering
    # This is MUCH faster than loading all 20k specimens and sorting in Python
    recent_specimens = (
        Specimen.objects
        .select_related('locality', 'recordedBy', 'georeferencedBy', 'identifiedBy')
        .order_by(
            F('eventDate').desc(nulls_last=True),
            F('eventTime').desc(nulls_last=True),
            '-year',
            '-catalogNumber'
        )[:20]
    )
    
    return render(request, 'butterflies/report_table.html', {
        'specimens': recent_specimens,
        'export_csv_url': reverse('export_report_csv'),
        'export_excel_url': reverse('export_report_excel'),
        # Not passing paginator or page_obj to template
    })

@guest_allowed
def guest_view(request):
    """
    Guest-friendly view that displays all filtered specimens in a scrollable list.
    Non-paginated for simplicity - shows all matching results.
    Optimized for performance with select_related and only() to minimize database queries.
    
    Parameters:
        request: HTTP request object
    Returns:
        Rendered template with filtered specimens
    """
    # Start with an optimized base queryset
    # Use select_related to avoid N+1 queries on foreign keys
    # Use only() to fetch only the fields we need for display
    # Handle view mode toggle (table/grid)
    view_mode = request.session.get('guest_view_mode', 'table')
    if request.method == 'POST' and 'toggle_view_mode' in request.POST:
        new_view_mode = request.POST.get('view_mode', 'table')
        request.session['guest_view_mode'] = new_view_mode
        request.session.modified = True
        return HttpResponseRedirect(request.get_full_path())

    specimens = Specimen.objects.select_related(
        'locality', 
        'recordedBy', 
        'georeferencedBy', 
        'identifiedBy'
    )
    
    # Define guest-specific special filters
    # Keep this minimal: map guest form fields to actual model/related fields
    special_filters = {
        # Locality hierarchy (guest labels → related fields)
        'department': {'field': 'locality__region', 'range_support': False},
        'province': {'field': 'locality__province', 'range_support': False},
        'district': {'field': 'locality__district', 'range_support': False},
        # Year supports ranges like 2020:2023 or comma-separated values
        'year': {'field': 'year', 'range_support': True},
    }
    
    # Determine if any guest filters were provided; if not, don't show results
    guest_filter_keys = [
        'family','genus','specificEpithet',
        'department','province','district',
        'minimumElevationInMeters','maximumElevationInMeters',
        'year','month','day'
    ]
    has_filters = any(request.GET.get(k, '').strip() for k in guest_filter_keys)

    if has_filters:
        # Apply all filters using the standardized filter utility
        specimens = apply_model_filters(specimens, Specimen, request, special_filters)
    else:
        specimens = Specimen.objects.none()
    
    # Fields for display and filtering
    fields = [
        {'name': field.name, 'verbose_name': getattr(field, 'verbose_name', field.name)}
        for field in Specimen._meta.fields
        if field.name != 'id'
    ]
    
    # Pagination (small initial payloads; faster perceived load)
    per_page = 50 if view_mode == 'table' else 40
    paginator = Paginator(specimens, per_page)
    page = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages) if paginator.num_pages else paginator.page(1)

    # Work only with current page for performance
    page_list = list(page_obj.object_list)

    # Add model_name_internal to each object for URL generation
    for specimen in page_list:
        specimen.model_name_internal = 'specimen'

    # If grid view, batch-load dorsal/ventral image URLs only for current page
    if view_mode == 'grid':
        catalog_numbers = [s.catalogNumber for s in page_list if s.catalogNumber]
        if catalog_numbers:
            batch_image_results = get_specimen_image_urls_batch(catalog_numbers)
        else:
            batch_image_results = {}
        for s in page_list:
            s.dorsal_image_url = None
            s.ventral_image_url = None
            if s.catalogNumber in batch_image_results:
                image_dict = batch_image_results[s.catalogNumber]
                if image_dict.get('dorsal') and image_dict['dorsal'] != 'no data':
                    s.dorsal_image_url = image_dict['dorsal']
                if image_dict.get('ventral') and image_dict['ventral'] != 'no data':
                    s.ventral_image_url = image_dict['ventral']

    # Total count for summary
    total_count = paginator.count

    # HTMX partial responses for infinite scroll
    is_htmx = request.headers.get('HX-Request', '').lower() == 'true'
    if is_htmx:
        template_name = (
            'butterflies/partials/_guest_grid_page.html'
            if view_mode == 'grid'
            else 'butterflies/partials/_guest_table_page.html'
        )
        return render(request, template_name, {
            'page_obj': page_obj,
            'view_mode': view_mode,
            'has_filters': has_filters,
        })

    # Full page response
    return render(request, 'butterflies/guest_view.html', {
        'fields': fields,
        'model_name': 'Specimen',
        'model_name_internal': 'specimen',
        'total_count': total_count,
        'request': request,
        'view_mode': view_mode,
        'supports_grid_view': True,
        'has_filters': has_filters,
        'page_obj': page_obj,
    })

# --- Export Views ---

def export_model_csv(request, model_name):
    """
    Generic function to export any model data as a CSV file.
    Optimized for large datasets using chunked processing.
    
    Parameters:
        request: HTTP request object
        model_name: String name of the model to export
    Returns:
        HttpResponse with CSV content for download
    """
    model = None
    for m in model_list():
        if m._meta.model_name == model_name:
            model = m
            break
    
    if not model:
        raise Http404("Model not found")
        
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{model_name}.csv"'
    writer = csv.writer(response)
    
    # Get exportable fields for this model (without related fields)
    exportable_fields = get_exportable_fields(model, include_related=False)
    headers = [field['name'] for field in exportable_fields]
    writer.writerow(headers)
    
    # Export all objects using iterator for memory efficiency
    CHUNK_SIZE = 1000
    for obj in model.objects.all().iterator(chunk_size=CHUNK_SIZE):
        row = []
        for field_info in exportable_fields:
            row.append(get_field_value_for_export(obj, field_info))
        writer.writerow(row)
    
    return response

def export_model_excel(request, model_name):
    """
    Generic function to export any model data as an Excel file.
    Optimized for large datasets using chunked processing.
    
    Parameters:
        request: HTTP request object
        model_name: String name of the model to export
    Returns:
        HttpResponse with Excel content for download
    """
    model = None
    for m in model_list():
        if m._meta.model_name == model_name:
            model = m
            break
    
    if not model:
        raise Http404("Model not found")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = model._meta.verbose_name_plural.title()
    
    # Get exportable fields for this model (without related fields)
    exportable_fields = get_exportable_fields(model, include_related=False)
    headers = [field['name'] for field in exportable_fields]
    ws.append(headers)
    
    # Export all objects in chunks to reduce memory usage
    CHUNK_SIZE = 1000
    batch_rows = []
    
    for obj in model.objects.all().iterator(chunk_size=CHUNK_SIZE):
        row = []
        for field_info in exportable_fields:
            row.append(get_field_value_for_export(obj, field_info))
        batch_rows.append(row)
        
        # Write in batches for better performance
        if len(batch_rows) >= CHUNK_SIZE:
            for batch_row in batch_rows:
                ws.append(batch_row)
            batch_rows = []
    
    # Write any remaining rows
    if batch_rows:
        for batch_row in batch_rows:
            ws.append(batch_row)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{model_name}.xlsx"'
    
    return response

# Helper function to get exportable fields for a model
def get_exportable_fields(model, include_related=True):
    """
    Get a list of field definitions for export.
    This function extracts all fields from a model and formats them for export.
    
    Args:
        model: Django model class
        include_related: Whether to include related fields from ForeignKeys
        
    Returns:
        List of dictionaries with field metadata
    """
    from django.db.models import ForeignKey
    exportable_fields = []
    
    # 1. Add direct fields from the model
    for field in model._meta.fields:
        if field.name != 'id':  # Skip the ID field
            exportable_fields.append({
                'name': field.name,
                'verbose_name': field.verbose_name,
                'is_relation': isinstance(field, ForeignKey),
                'related_model': field.related_model.__name__ if isinstance(field, ForeignKey) else None
            })
    
    # 2. Add related fields if requested
    if include_related and model == Specimen:
        # Add locality-related fields
        locality_fields = [
            ('country', 'country'),
            ('stateProvince', 'province'),
            ('county', 'region'),
            ('municipality', 'district'),
            ('habitat', 'habitat')
        ]
        
        for export_name, django_field in locality_fields:
            exportable_fields.append({
                'name': export_name,
                'verbose_name': export_name,
                'is_relation': True,
                'related_model': 'Locality',
                'related_field': django_field
            })
    
    return exportable_fields

# Helper function to extract a field's value
def get_field_value_for_export(obj, field_info):
    """
    Extract a field value from an object based on field info.
    Handles regular fields, foreign keys, and related fields.
    
    Args:
        obj: Model instance
        field_info: Dictionary with field metadata
        
    Returns:
        Field value formatted for export
    """
    if field_info.get('is_relation') and field_info.get('related_model') == 'Locality':
        # Handle locality-related fields
        locality = getattr(obj, 'locality', None)
        if not locality:
            return ''
        return getattr(locality, field_info.get('related_field', ''), '')
    elif field_info.get('is_relation'):
        # Handle other foreign key fields
        related_obj = getattr(obj, field_info['name'], None)
        return str(related_obj) if related_obj else ''
    else:
        # Handle direct fields
        return getattr(obj, field_info['name'], '')

def export_report_csv(request):
    """
    Export the full report table (specimens with all related fields) as a CSV file.
    Optimized for large datasets using chunked processing.
    
    Parameters:
        request: HTTP request object
    Returns:
        HttpResponse with CSV content for download
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="specimen_report.csv"'
    writer = csv.writer(response)
    
    # Get exportable fields
    exportable_fields = get_exportable_fields(Specimen)
    
    # Build headers from exportable fields
    headers = [field['name'] for field in exportable_fields]
    writer.writerow(headers)
    
    # Get all specimens with related data using iterator for memory efficiency
    CHUNK_SIZE = 1000
    specimens = Specimen.objects.select_related(
        'locality', 'recordedBy', 'georeferencedBy', 'identifiedBy'
    ).iterator(chunk_size=CHUNK_SIZE)
    
    # Write each specimen row with all fields
    for specimen in specimens:
        row = []
        for field_info in exportable_fields:
            row.append(get_field_value_for_export(specimen, field_info))
        writer.writerow(row)
    
    return response

def export_report_excel(request):
    """
    Export the full report table (specimens with all related fields) as an Excel file.
    Optimized for large datasets using chunked processing and efficient Excel writing.
    
    Parameters:
        request: HTTP request object
    Returns:
        HttpResponse with Excel content for download
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Specimen Report"
    
    # Get exportable fields
    exportable_fields = get_exportable_fields(Specimen)
    
    # Build headers from exportable fields
    headers = [field['name'] for field in exportable_fields]
    ws.append(headers)
    
    # Process specimens in chunks to reduce memory usage
    CHUNK_SIZE = 1000
    batch_rows = []
    
    # Get all specimens with related data using iterator for memory efficiency
    specimens = Specimen.objects.select_related(
        'locality', 'recordedBy', 'georeferencedBy', 'identifiedBy'
    ).iterator(chunk_size=CHUNK_SIZE)
    
    # Write each specimen row with all fields
    for specimen in specimens:
        row = []
        for field_info in exportable_fields:
            row.append(get_field_value_for_export(specimen, field_info))
        batch_rows.append(row)
        
        # Write in batches for better performance
        if len(batch_rows) >= CHUNK_SIZE:
            for batch_row in batch_rows:
                ws.append(batch_row)
            batch_rows = []
    
    # Write any remaining rows
    if batch_rows:
        for batch_row in batch_rows:
            ws.append(batch_row)
    
    # Simplified column width adjustment (removed expensive iteration)
    # Set reasonable default widths instead of calculating from all cells
    for i, field_info in enumerate(exportable_fields, start=1):
        column_letter = openpyxl.utils.get_column_letter(i)
        # Set width based on header length with reasonable defaults
        header_length = len(field_info['name'])
        ws.column_dimensions[column_letter].width = min(max(header_length + 2, 15), 50)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="specimen_report.xlsx"'
    
    return response

# --- Import Views ---

# Helper function to validate and construct eventDate from components
def validate_and_construct_event_date(day, month, year, debug_mode=False):
    """
    Validates day, month, year components and attempts to construct a valid date.
    
    Returns:
        tuple: (date_obj, error_message, is_valid)
            - date_obj: datetime.date or None if invalid
            - error_message: String with error details or None if valid
            - is_valid: Boolean indicating if the date is valid
    """
    date_obj = None
    date_error = None
    
    # If not all components are provided, we can't validate
    if not (day and month and year):
        missing = []
        if not day: missing.append('day')
        if not month: missing.append('month')
        if not year: missing.append('year')
        components_str = ', '.join(missing)
        return None, f"Missing date components: {components_str}", False
    
    try:
        # Try numeric month parsing first
        if month.isdigit() and 1 <= int(month) <= 12:
            try:
                date_obj = datetime.date(int(year), int(month), int(day))
            except ValueError as e:
                date_error = f"Invalid date: {e}"
        else:
            # Try abbreviated and full month names
            temp_date_str = f"{year}-{month}-{day}"
            for fmt in ["%Y-%b-%d", "%Y-%B-%d"]:
                try:
                    date_obj = datetime.datetime.strptime(temp_date_str, fmt).date()
                    break
                except ValueError:
                    continue
        
        # If direct parsing failed, try string parsing as fallback
        if not date_obj:
            date_str = f"{day} {month}, {year}"
            try:
                parsed_date = parse_date_value(date_str)
                if parsed_date:
                    date_obj = parsed_date
            except Exception as e:
                if not date_error:
                    date_error = f"Could not parse date from components: {e}"
        
        # Return the results
        if date_obj:
            return date_obj, None, True
        else:
            error = date_error or "Could not construct a valid date from day, month, and year"
            return None, error, False
            
    except Exception as e:
        return None, f"Error validating date components: {str(e)}", False
        
# Helper function for processing date fields in a unified way
def process_date_fields_unified(row_data, request=None, row_index=None, context=None, debug_mode=False):
    """
    Process all date fields in a row, converting string dates to date objects.
    Can be used both for preview and import phases.
    
    Args:
        row_data (dict): The row data containing date fields
        request (HttpRequest): Optional request object for adding messages (import mode)
        row_index (int): Optional index of the current row for error messages (import mode)
        context (dict): Optional preview context for adding warnings/errors (preview mode)
        debug_mode (bool): Whether to enable additional debug messages
        
    Returns:
        dict: Dictionary of parse results by field (field_name -> date_obj) 
              or None if used in preview mode
    """
    # Define date fields to process
    date_fields = ['eventDate', 'dateIdentified', 'georeferencedDate']
    results = {}
    
    # Process each date field
    for field in date_fields:
        if field in row_data and row_data[field]:
            # Skip eventDate special handling for now - will be processed separately
            if field == 'eventDate':
                continue
                
            # Use our helper function to parse date fields
            date_obj, error_msg, is_valid = parse_date_field(row_data[field], debug_mode)
            
            if is_valid and date_obj:
                # Store the date object
                row_data[field] = date_obj
                results[field] = date_obj
                
                # Add info message in different modes
                if debug_mode:
                    if request and hasattr(request, '_import_context'):  # Import mode
                        messages.info(request, f"Converted date for {field} from '{row_data[field]}' to '{date_obj}'")
                    elif context:  # Preview mode
                        formatted = format_date_value(date_obj)
                        context['warnings'].append(
                            f"Date '{row_data[field]}' will be stored as '{date_obj}' and displayed as '{formatted}'"
                        )
            else:
                # Handle invalid date
                row_data[field] = None
                results[field] = None
                
                if debug_mode:
                    if request and hasattr(request, '_import_context'):  # Import mode
                        messages.warning(request, f"{error_msg} for {field}, set to None")
                    elif context:  # Preview mode
                        context['errors'].append(f"{error_msg} for {field}")
                        context['error_fields'].add(field)
    
    # Handle eventDate separately using the unified process_event_date function
    success, skip_row, event_date_obj = process_event_date(row_data, context, request, row_index, debug_mode)
    if event_date_obj:
        results['eventDate'] = event_date_obj
    
    return results if not context else None
    
# Helper function for processing foreign keys during import
def process_foreign_keys(row_data, row_index, request, debug_mode=False):
    """
    Process foreign key relationships for a specimen row.
    
    Args:
        row_data (dict): The row data containing foreign key fields
        row_index (int): The index of the current row for error messages
        request (HttpRequest): The request object for adding messages
        debug_mode (bool): Whether to enable additional debug messages
    """
    # Define foreign key fields and their related models
    fk_fields = {
        'locality': (Locality, 'localityCode'),
        'recordedBy': (Initials, 'initials'),
        'georeferencedBy': (Initials, 'initials'),
        'identifiedBy': (Initials, 'initials')
    }
    
    # Process each foreign key field
    for fk_field, (related_model, lookup_field) in fk_fields.items():
        if fk_field in row_data and row_data[fk_field]:
            lookup_value = row_data[fk_field]
            
            # Special case: allow "." as a placeholder for identifiedBy (treat as None/NULL)
            if fk_field == 'identifiedBy' and lookup_value == '.':
                if debug_mode and request and hasattr(request, '_import_context'):
                    messages.info(
                        request,
                        f"Found '.' placeholder for {fk_field} in row {row_index+1}. Setting to NULL."
                    )
                row_data[fk_field] = None
                continue
                
            # Normal case: look up the related object
            related_obj = related_model.objects.filter(**{lookup_field: lookup_value}).first()
            if related_obj:
                row_data[fk_field] = related_obj
            else:
                row_data[fk_field] = None
                messages.warning(
                    request, 
                    f"{related_model.__name__} '{lookup_value}' not found for {fk_field} in row {row_index+1}."
                )


# Helper function to parse a date field
def parse_date_field(date_value, debug_mode=False):
    """
    Attempts to parse a date string into a datetime.date object.
    
    Returns:
        tuple: (date_obj, error_message, is_valid)
            - date_obj: datetime.date or None if invalid
            - error_message: String with error details or None if valid
            - is_valid: Boolean indicating if the date is valid
    """
    if not date_value or date_value == '.':
        return None, None, True  # Empty or placeholder, not an error
    
    # Convert to string if not already
    if not isinstance(date_value, str):
        date_value = str(date_value)
    
    # If it has a time component (contains space), extract just the date part
    if ' ' in date_value:
        try:
            date_value = date_value.split(' ')[0]
        except Exception:
            pass
        
    try:
        # Try direct parsing with dateutil
        try:
            date_obj = date_parser.parse(date_value).date()
            return date_obj, None, True
        except Exception as dateutil_error:
            # If direct parsing fails, try our standard format
            try:
                date_obj = parse_date_value(date_value)
                if date_obj:
                    return date_obj, None, True
            except Exception:
                pass
            
            # Try ISO format specifically
            try:
                # Handle potential timezone information
                if 'T' in date_value:
                    date_value = date_value.split('T')[0]
                date_obj = datetime.date.fromisoformat(date_value)
                return date_obj, None, True
            except ValueError:
                pass
                
        # If we get here, all parsing attempts failed
        return None, f"Could not parse date '{date_value}'", False
    except Exception as e:
        return None, f"Invalid date format '{date_value}': {str(e)}", False
        

# (The process_date_fields function has been replaced by process_date_fields_unified)


# Helper function to build the cache of valid foreign key values for specimen validation
def build_fk_validation_cache():
    """
    Builds a cache of valid foreign key values for specimen validation.
    
    Returns:
        tuple: (fk_fields_validation, valid_values_cache)
            - fk_fields_validation: Dictionary mapping foreign key fields to their related models and fields
            - valid_values_cache: Dictionary mapping foreign key fields to sets of valid values
    """
    fk_fields_validation = {
        'locality': (Locality, 'localityCode', 'Locality'),
        'recordedBy': (Initials, 'initials', 'Initials'),
        'georeferencedBy': (Initials, 'initials', 'Initials'),
        'identifiedBy': (Initials, 'initials', 'Initials')
    }
    
    # Build a cache of valid values
    valid_values_cache = {}
    for fk_field, (related_model, lookup_field, display_name) in fk_fields_validation.items():
        valid_values_cache[fk_field] = set(
            related_model.objects.values_list(lookup_field, flat=True)
        )
    
    return fk_fields_validation, valid_values_cache
    
# Helper function for managing import errors
def handle_import_error(request, error_msg, row_index):
    """
    Add an import error to the session for display later.
    
    Args:
        request (HttpRequest): The request object
        error_msg (str): The error message
        row_index (int): The index of the row with the error
    """
    user_message = f"Error in row {row_index+1}: {error_msg}"
    messages.error(request, user_message)
    import_errors = request.session.get('import_errors', [])
    import_errors.append(user_message)
    request.session['import_errors'] = import_errors

# Helper function for validating specimen data - unified approach for preview and import
def validate_specimen_data(row_data, preview_item=None, common_errors=None, debug_mode=False):
    """
    Performs common validation for specimen data.
    
    Parameters:
        row_data: Dictionary containing the specimen data
        preview_item: Optional dictionary with preview metadata to update
                     If None, only returns errors without modifying preview_item
        common_errors: Dictionary of pre-identified common errors (optional)
        debug_mode: Boolean indicating whether to show detailed debug information
    
    Returns:
        Dictionary of errors by field (if preview_item is None)
        Otherwise, updates preview_item in place and returns None
    """
    # Use for collecting errors if preview_item is None
    collected_errors = {}
    collected_warnings = []
    collected_error_fields = set()
    
    # Get the foreign key fields validation info
    fk_fields_validation, valid_values_cache = build_fk_validation_cache()
    
    # Validate foreign key fields
    for fk_field, (related_model, lookup_field, display_name) in fk_fields_validation.items():
        if fk_field in row_data and row_data[fk_field]:
            lookup_value = row_data[fk_field]
            
            # Special case: allow "." as a placeholder for identifiedBy
            if fk_field == 'identifiedBy' and lookup_value == '.':
                # This is valid - it will be converted to NULL/None during import
                continue
            
            key = f"{fk_field}_invalid"
            
            # Check if this value is invalid using the common_errors or validation cache
            is_invalid = False
            
            # Use common_errors if available (faster than DB check)
            if common_errors and key in common_errors and lookup_value in common_errors[key]:
                is_invalid = True
                
            # Also check validation cache
            elif lookup_value not in valid_values_cache[fk_field]:
                is_invalid = True
            
            if is_invalid:
                error_msg = f"Invalid {display_name}: '{lookup_value}' does not exist in the database"
                if preview_item:
                    preview_item['errors'].append(error_msg)
                    preview_item['error_fields'].add(fk_field)
                else:
                    collected_errors[fk_field] = error_msg
                    collected_error_fields.add(fk_field)
    
    # Validate sex field
    if 'sex' in row_data and row_data['sex']:
        sex_invalid = False
        
        # First check common errors if available
        if common_errors and 'sex_values' in common_errors and row_data['sex'] in common_errors['sex_values']:
            sex_invalid = True
        # Always validate directly too
        elif row_data['sex'] not in ['male', 'female', '.']:
            sex_invalid = True
            
        if sex_invalid:
            error_msg = f"Invalid sex: '{row_data['sex']}'. Must be 'male', 'female', or '.'"
            if preview_item:
                preview_item['errors'].append(error_msg)
                preview_item['error_fields'].add('sex')
            else:
                collected_errors['sex'] = error_msg
                collected_error_fields.add('sex')
    
    # Validate dates
    date_fields = ['eventDate', 'dateIdentified', 'georeferencedDate']
    for date_field in date_fields:
        if date_field in row_data and row_data[date_field]:
            original_value = row_data[date_field]
            # Skip the "." placeholder, as it will be filled from components
            if original_value == '.':
                continue
            
            # Use our helper function to validate the date
            date_obj, error_msg, is_valid = parse_date_field(original_value, debug_mode)
            
            if is_valid and date_obj:
                # Only show conversion information in debug mode
                if debug_mode:
                    formatted = format_date_value(date_obj)
                    warning_msg = f"Date '{original_value}' will be stored as '{date_obj}' and displayed as '{formatted}'"
                    if preview_item:
                        preview_item['warnings'].append(warning_msg)
                    else:
                        collected_warnings.append(warning_msg)
            elif not is_valid:
                # Only show parsing failures as errors, not warnings
                error_msg = f"{error_msg} for {date_field}"
                if preview_item:
                    preview_item['errors'].append(error_msg)
                    preview_item['error_fields'].add(date_field)
                else:
                    collected_errors[date_field] = error_msg
                    collected_error_fields.add(date_field)
    
    # Check catalogNumber components
    if not row_data.get('catalogNumber'):
        missing_components = []
        
        if not row_data.get('specimenNumber'):
            missing_components.append('specimenNumber')
        
        if not row_data.get('year'):
            missing_components.append('year')
        
        if 'locality' not in row_data or not row_data['locality']:
            missing_components.append('locality')
        
        if missing_components:
            components_str = ', '.join(missing_components)
            error_msg = f"Missing required field(s): {components_str}. These are needed to generate catalogNumber."
            
            if preview_item:
                preview_item['errors'].append(error_msg)
                for field in missing_components:
                    preview_item['error_fields'].add(field)
            else:
                for field in missing_components:
                    collected_errors[field] = "Required for catalogNumber"
                    collected_error_fields.add(field)
    
    # Check boolean fields
    if 'exact_loc' in row_data and row_data['exact_loc']:
        value = str(row_data['exact_loc']).upper()
        if value not in ['TRUE', 'FALSE']:
            warning_msg = f"Value '{row_data['exact_loc']}' for exact_loc will be converted to TRUE/FALSE"
            if preview_item:
                preview_item['warnings'].append(warning_msg)
            else:
                collected_warnings.append(warning_msg)
    
    # Process eventTime to strip seconds (HH:MM:SS → HH:MM)
    if 'eventTime' in row_data and row_data['eventTime']:
        time_value = str(row_data['eventTime'])
        # Only strip seconds if it's a standard time format (not a range with hyphens or other special formats)
        if time_value.count(':') == 2 and '-' not in time_value and '/' not in time_value and ',' not in time_value:
            # Extract just the hours and minutes (HH:MM)
            try:
                hour_minute = ':'.join(time_value.split(':')[:2])
                row_data['eventTime'] = hour_minute
                
                # Add info message
                warning_msg = f"Stripped seconds from time value: '{time_value}' → '{hour_minute}'"
                if preview_item:
                    preview_item['warnings'].append(warning_msg)
                else:
                    collected_warnings.append(warning_msg)
            except Exception:
                # Continue with original value if stripping fails
                pass
                
    # Return collected errors if no preview_item was provided
    if not preview_item:
        return {
            'errors': collected_errors,
            'warnings': collected_warnings,
            'error_fields': collected_error_fields
        }
    
    return None

# Helper function for processing eventDate from components - unified for preview and import
def process_event_date(row_data, context=None, request=None, row_index=None, debug_mode=False):
    """
    Validates and processes eventDate from day, month, year components.
    Can be used both for preview display and for actual import.
    
    Parameters:
        row_data: Dictionary containing row data with day, month, year components
        context: Optional dictionary for preview metadata (warnings/errors)
        request: Optional HttpRequest object for import errors
        row_index: Optional row index for error messages during import
        debug_mode: Boolean indicating whether to show detailed debug information
        
    Returns:
        tuple: (success, skip_row, date_obj)
            - success: Boolean indicating if the date was successfully processed
            - skip_row: Boolean indicating if the row should be skipped (import only)
            - date_obj: The date object if successfully created, None otherwise
    """
    # Only process if eventDate is missing or "."
    if not row_data.get('eventDate') or row_data.get('eventDate') == '.':
        day = row_data.get('day')
        month = row_data.get('month')
        year = row_data.get('year')
        
        # Use our existing function to validate and construct the date
        date_obj, error_msg, is_valid = validate_and_construct_event_date(day, month, year, debug_mode)
        
        # At least one component exists but validation failed
        if (day or month or year) and not is_valid:
            # Handle validation failure differently for preview vs import
            if context:  # Preview mode
                context['errors'].append(error_msg)
                
                # Mark relevant fields as having errors
                if not day: context['error_fields'].add('day')
                if not month: context['error_fields'].add('month')
                if not year: context['error_fields'].add('year')
                context['error_fields'].add('eventDate')
                
                return False, False, None
            elif request:  # Import mode
                handle_import_error(request, error_msg, row_index)
                
                # If in debug mode, continue with None value rather than failing
                if debug_mode and hasattr(request, '_import_context'):
                    messages.warning(request, f"Row {row_index+1}: Using NULL for eventDate due to parsing error")
                    row_data['eventDate'] = None
                    return True, False, None
                else:
                    return False, True, None
            else:
                # Basic mode just returns the status
                return False, False, None
        
        # All components exist and validation succeeded
        elif is_valid and date_obj:
            # Handle success differently for preview vs import
            if context:  # Preview mode
                # Success! We'll add an informational message
                context['warnings'].append(
                    f"Will generate eventDate from components: {day} {month} {year} → {date_obj}"
                )
                # Update the preview data to show the generated eventDate
                row_data['eventDate'] = str(date_obj)
                context['data']['eventDate'] = str(date_obj)
                
                return True, False, date_obj
            else:  # Import mode or basic mode
                row_data['eventDate'] = date_obj
                
                # Add debug message if in import mode
                if request and debug_mode and hasattr(request, '_import_context'):
                    messages.info(request, f"Generated eventDate '{format_date_value(date_obj)}' from components: {day} {month} {year}")
                
                return True, False, date_obj
        
        # Have some but not all components
        elif (day or month or year) and not (day and month and year):
            error_msg = f"Incomplete date components: day={day or 'missing'}, month={month or 'missing'}, year={year or 'missing'}"
            
            if context:  # Preview mode
                context['errors'].append(error_msg)
                # Mark fields as having errors
                if not day: context['error_fields'].add('day')
                if not month: context['error_fields'].add('month')
                if not year: context['error_fields'].add('year')
                context['error_fields'].add('eventDate')
                
                return False, False, None
            elif request:  # Import mode
                handle_import_error(request, error_msg, row_index)
                
                # If in debug mode, continue with None value rather than failing
                if debug_mode and hasattr(request, '_import_context'):
                    messages.warning(request, f"Row {row_index+1}: Using NULL for eventDate due to incomplete components")
                    row_data['eventDate'] = None
                    return True, False, None
                else:
                    return False, True, None
            else:
                # Basic mode just returns the status
                return False, False, None
            
    return True, False, None

@csrf_exempt
@admin_required
def import_model(request, model_name):
    """
    Generic function to import data for any model from CSV or Excel file.
    This simplified version focuses on:
    1. Reading all values as text without restrictions
    2. Converting data to proper formats before import
    3. Showing a clear preview with potential issues
    4. Handling imports with better error messages
    
    Parameters:
        request: HTTP request object
        model_name: String name of the model to import
    Returns:
        Rendered template for import form, preview, or redirect
    """
    # Mark this request as import context for debug message filtering
    request._import_context = True
    
    # Using the helper functions defined outside this function
    
    # DEBUG: Log which model we're importing
    print(f"\n\n========== DEBUGGING IMPORT PROCESS ==========")
    print(f"Import requested for model: {model_name}")
    import inspect
    print(f"Called from: {inspect.stack()[1].function}")
    print(f"Request method: {request.method}")
    if request.method == 'POST':
        print(f"POST data keys: {list(request.POST.keys())}")
        print(f"FILES: {list(request.FILES.keys()) if request.FILES else 'No files'}")
    print(f"================================================\n\n")

    # Find the model
    model = None
    print(f"\n[DEBUG] Searching for model: {model_name}")
    available_models = []
    for m in model_list():
        available_models.append(m._meta.model_name)
        if m._meta.model_name == model_name:
            model = m
            break
    
    print(f"[DEBUG] Available models: {available_models}")
    print(f"[DEBUG] Found model: {model._meta.model_name if model else 'None'}")
    
    if not model:
        print(f"[ERROR] Model '{model_name}' not found in available models")
        raise Http404("Model not found")
    
    # Prepare context with model information
    context = {
        'model_name': model._meta.verbose_name,
        'model_name_plural': model._meta.verbose_name_plural,
        'model_name_internal': model_name,
        'model_fields': [f.name for f in model._meta.fields if f.name != 'id'],
    }
    
    # For Specimen model, we only check catalogNumber for uniqueness
    # For other models, find unique field for duplicate detection
    unique_field = None
    if model_name == 'specimen':
        unique_field = 'catalogNumber'  # Specifically use catalogNumber for specimens
    else:
        # For other models, use the first unique field found (excluding id)
        for field in model._meta.fields:
            if field.unique and field.name != 'id':
                unique_field = field.name
                break
            
        # Check unique_together constraints if no unique field found
        if not unique_field:
            unique_together = getattr(model._meta, 'unique_together', [])
            if unique_together and isinstance(unique_together[0], (list, tuple)) and len(unique_together[0]) > 0:
                unique_field = unique_together[0][0]
    
    # Add unique field info to context
    context['has_unique_field'] = bool(unique_field)
    context['unique_field'] = unique_field
    
    # Reset any previous import session data
    if 'has_errors' in request.session:
        del request.session['has_errors']
    
    # Step 1: Handle file upload and parse into DataFrame
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        file_ext = file.name.split('.')[-1].lower()
        
        print(f"\n[DEBUG] Processing file: {file.name} ({file_ext})")
        print(f"[DEBUG] File size: {file.size} bytes")
        
        try:
            # Read file contents - all values as strings initially
            if file_ext == 'csv':
                # Use dtype=str to read everything as text
                print(f"[DEBUG] Reading CSV file")
                
                # First determine the number of non-empty rows
                print(f"[DEBUG] Scanning CSV file for non-empty rows")
                try:
                    # Create a copy of the file in memory to avoid consuming the file object
                    file_copy = io.StringIO(file.read().decode('utf-8'))
                    file.seek(0)  # Reset the file pointer for later reading
                    
                    # Sample the first 100 rows to get column count
                    sample_df = pd.read_csv(file_copy, nrows=100, dtype=str)
                    file_copy.seek(0)  # Reset for full scan
                    
                    # Count non-empty rows efficiently using csv module
                    csv_reader = csv.reader(file_copy)
                    headers = next(csv_reader)  # Get headers
                    
                    # Count rows that have at least one non-empty field
                    row_count = 0
                    max_rows = 1000000  # Safety limit to prevent infinite loops
                    for i, row in enumerate(csv_reader):
                        if i >= max_rows:
                            print(f"[WARNING] Reached safety limit of {max_rows} rows")
                            break
                        
                        if any(field.strip() for field in row):
                            row_count += 1
                    
                    print(f"[DEBUG] Found {row_count} non-empty rows in CSV (plus header)")
                    
                    # Read only the non-empty rows (plus some buffer)
                    nrows = min(row_count + 100, max_rows)  # Add buffer for safety
                    df = pd.read_csv(file, dtype=str, keep_default_na=False, nrows=nrows)
                    
                except Exception as e:
                    print(f"[DEBUG] Error in CSV row scanning: {str(e)}, falling back to standard read")
                    file.seek(0)  # Reset file pointer
                    df = pd.read_csv(file, dtype=str, keep_default_na=False)
                
            elif file_ext in ('xls', 'xlsx'):
                # Use dtype=str to read everything as text initially
                print(f"[DEBUG] Reading Excel file")
                
                try:                
                    # Save the uploaded file to a temporary file
                    with tempfile.NamedTemporaryFile(delete=False, suffix=f".{file_ext}") as temp_file:
                        for chunk in file.chunks():
                            temp_file.write(chunk)
                        temp_path = temp_file.name
                    
                    # Open the workbook and get the first sheet
                    print(f"[DEBUG] Opening Excel workbook to determine data range")
                    wb = openpyxl.load_workbook(temp_path, read_only=True, data_only=True)
                    sheet = wb.active
                    
                    # Find the actual data range
                    max_row = 0
                    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row), 1):
                        if any(cell.value for cell in row):
                            max_row = row_idx
                    
                    # Add some buffer rows
                    max_row = min(max_row + 10, 1000000)  # Safety limit
                    print(f"[DEBUG] Determined Excel data extends to row {max_row}")
                    
                    # Close the workbook and clean up
                    wb.close()
                    
                    # Now read only the rows with data
                    file.seek(0)  # Reset file pointer
                    df = pd.read_excel(file, dtype=str, nrows=max_row)
                    
                    # Remove the temporary file
                    os.unlink(temp_path)
                    
                except Exception as e:
                    print(f"[DEBUG] Error in Excel row scanning: {str(e)}, falling back to standard read")
                    file.seek(0)  # Reset file pointer
                    df = pd.read_excel(file, dtype=str)
                
                # Convert all numeric columns to string to prevent automatic float conversion
                for col in df.columns:
                    df[col] = df[col].astype(str)
            else:
                print(f"[ERROR] Unsupported file type: {file_ext}")
                messages.error(request, 'Unsupported file type. Please upload a CSV or Excel file.')
                context['error'] = 'Unsupported file type. Please upload a CSV or Excel file.'
                return render(request, 'butterflies/import_model.html', context)
            
            print(f"[DEBUG] DataFrame loaded, shape: {df.shape}")
            print(f"[DEBUG] DataFrame columns: {list(df.columns)}")
            print(f"[DEBUG] First 2 rows: {df.head(2).to_dict(orient='records')}")
            
            # Basic cleanup - strip whitespace from all string values
            print(f"[DEBUG] Starting data cleanup")
            for col in df.columns:
                df[col] = df[col].astype(str).str.strip()
            
            # Replace empty strings with None more efficiently
            print(f"[DEBUG] Replacing empty values with None")
            empty_values = ['', 'nan', 'None', 'NaN', 'null', 'NULL']
            
            # Process column by column to be more memory efficient
            for col in df.columns:
                df[col] = df[col].apply(lambda x: None if x in empty_values or pd.isna(x) or x is None else x)
            
            # Remove completely empty rows more efficiently
            print(f"[DEBUG] Removing empty rows")
            original_rows = len(df)
            
            # Create a mask of rows that have at least one non-None value
            # This is more efficient than dropna for large dataframes
            non_empty_mask = df.applymap(lambda x: x is not None and x != '').any(axis=1)
            df = df[non_empty_mask]
            
            print(f"[DEBUG] Removed {original_rows - len(df)} empty rows")
            
            # Check if dataframe is empty after cleaning
            if df.empty:
                print(f"[ERROR] DataFrame is empty after cleaning")
                messages.error(request, 'The uploaded file contains no usable data after cleaning.')
                context['error'] = 'The uploaded file contains no usable data after cleaning.'
                return render(request, 'butterflies/import_model.html', context)
            
            print(f"[DEBUG] After cleanup, DataFrame shape: {df.shape}")
            print(f"[DEBUG] After cleanup, first row: {df.iloc[0].to_dict()}")
                
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
            context['error'] = f'Error processing file: {str(e)}'
            return render(request, 'butterflies/import_model.html', context)
        
        # Step 2: Check columns and prepare data for preview
        # Get expected field names
        expected_fields = [f.name for f in model._meta.fields if f.name != 'id']
        print(f"\n[DEBUG] Expected fields for {model_name}: {expected_fields}")
        
        # Map any foreign key fields to their expected column names for import
        if model_name == 'specimen':
            fk_fields = {
                'locality': 'locality',        # Locality.localityCode
                'recordedBy': 'recordedBy',     # Initials.initials
                'georeferencedBy': 'georeferencedBy', # Initials.initials
                'identifiedBy': 'identifiedBy',   # Initials.initials
            }
            print(f"[DEBUG] Using foreign key mappings for specimen: {fk_fields}")
        else:
            fk_fields = {}
            print(f"[DEBUG] No foreign key mappings for {model_name}")
        
        # Check for missing columns
        available_columns = list(df.columns)
        missing_columns = []
        
        print(f"[DEBUG] Available columns in file: {available_columns}")
        
        for field in expected_fields:
            # Check if field is in DataFrame columns or has a mapped name
            if field not in available_columns and field not in fk_fields.keys():
                missing_columns.append(field)
                print(f"[DEBUG] Missing column: {field}")
        
        if missing_columns:
            context['error'] = f'Missing columns: {", ".join(missing_columns)}'
            context['available_columns'] = available_columns
            context['expected_columns'] = expected_fields
            return render(request, 'butterflies/import_model.html', context)
        
        # Store raw data in session for pagination
        # We store the raw DataFrame to avoid re-processing every time
        request.session['import_df'] = serialize_dataframe(df)
        request.session['import_expected_fields'] = expected_fields
        request.session['import_fk_fields'] = fk_fields
        request.session['import_unique_field'] = unique_field
        request.session['import_model_name'] = model_name
        request.session['import_all_errors'] = False  # Reset error tracking
        request.session['show_errors_only'] = False  # Reset filter state for new import
        request.session.modified = True
        
        # Set up pagination
        page = 1  # Start with first page
        rows_per_page = 20  # Set a reasonable default
        
        # Now process paginated data
        return process_paginated_import_data(request, page, rows_per_page, context)
    
    # Handle pagination navigation
    elif request.method == 'POST' and request.POST.get('page'):
        # Get pagination parameters
        page = int(request.POST.get('page', 1))
        rows_per_page = int(request.POST.get('rows_per_page', 20))
        
        # Check if we have stored data
        if 'import_df' not in request.session:
            messages.error(request, 'Session expired or invalid data. Please upload the file again.')
            return redirect('import_model', model_name=model_name)
        
        # Reset error filter if we're navigating and it was automatically disabled
        if request.session.get('show_errors_only', False):
            # Quick check if there are any errors in the dataset at all
            df = deserialize_dataframe(request.session.get('import_df'))
            if df is not None:
                # Simple check - if we're still in error-only mode, validate it's still needed
                # This will be properly handled in process_paginated_import_data
                pass
        
        # Process the requested page
        return process_paginated_import_data(request, page, rows_per_page, context)
    
    # Handle error filter toggle
    elif request.method == 'POST' and (request.POST.get('show_errors_only') or request.POST.get('show_all')):
        # Get current pagination parameters
        page = int(request.POST.get('current_page', 1))
        rows_per_page = int(request.POST.get('rows_per_page', 20))
        
        # Check if we have stored data
        if 'import_df' not in request.session:
            messages.error(request, 'Session expired or invalid data. Please upload the file again.')
            return redirect('import_model', model_name=model_name)
        
        # Toggle the error filter state
        if request.POST.get('show_errors_only'):
            request.session['show_errors_only'] = True
            messages.info(request, 'Now showing only rows with errors.')
        else:
            request.session['show_errors_only'] = False
            messages.info(request, 'Now showing all rows.')
        
        request.session.modified = True
        
        # Process the current page with the new filter
        return process_paginated_import_data(request, page, rows_per_page, context)
    
    # Handle applying edits (save edited values)
    elif request.method == 'POST' and request.POST.get('apply_edits'):
        # Get pagination parameters
        page = int(request.POST.get('current_page', 1))
        rows_per_page = int(request.POST.get('rows_per_page', 20))
        
        # Check if we have stored data
        if 'import_df' not in request.session:
            messages.error(request, 'Session expired or invalid data. Please upload the file again.')
            return redirect('import_model', model_name=model_name)
        
        # Apply the edits and reprocess the current page
        df = deserialize_dataframe(request.session.get('import_df'))
        if df is None:
            messages.error(request, 'Session expired or invalid data. Please upload the file again.')
            return redirect('import_model', model_name=model_name)
        
        # Apply edited values to the DataFrame
        total_rows = len(df)
        start_idx = (page - 1) * rows_per_page
        end_idx = min(start_idx + rows_per_page, total_rows)
        
        fields = request.session.get('import_expected_fields', [])
        expected_fields = fields  # Make sure expected_fields is available for revalidation
        fk_fields = request.session.get('import_fk_fields', {})
        edited_count = 0
        edited_rows = set()  # Track which rows were actually edited
        
        # Only process rows that actually have form data (important when error filter is active)
        for i, idx in enumerate(range(start_idx, end_idx)):
            # Check if this row has any form data - if not, skip it entirely
            has_form_data = False
            for field in fields:
                form_key = f'row_{i}_{field}'
                if form_key in request.POST:
                    has_form_data = True
                    break
            
            # Skip this row if no form data found (means it wasn't visible in the form)
            if not has_form_data:
                continue
            
            row_was_edited = False
            for field in fields:
                form_key = f'row_{i}_{field}'
                # Only process if the form field actually exists in POST data
                if form_key not in request.POST:
                    continue
                    
                new_value = request.POST.get(form_key, '').strip()
                if new_value == '':
                    new_value = None
                
                # Get current value for comparison
                current_value = None
                if field in fk_fields:
                    col_name = fk_fields[field]
                    if col_name in df.columns:
                        current_value = df.at[idx, col_name]
                elif field in df.columns:
                    current_value = df.at[idx, field]
                
                # Convert current_value to comparable format
                if current_value is not None:
                    current_value = str(current_value).strip()
                    if current_value == '':
                        current_value = None
                
                # Check if value actually changed
                if new_value != current_value:
                    edited_count += 1
                    row_was_edited = True
                    # Update the DataFrame
                    if field in fk_fields:
                        col_name = fk_fields[field]
                        df.at[idx, col_name] = new_value
                    elif field in df.columns:
                        df.at[idx, field] = new_value
            
            # Track this row as edited if any field was changed
            if row_was_edited:
                edited_rows.add(idx)
        
        # Save updated dataframe back to session
        request.session['import_df'] = serialize_dataframe(df)
        request.session.modified = True
        
        # If any rows were edited, revalidate only those rows efficiently
        fixed_errors = 0
        if edited_rows:
            # Get validation setup for efficient checking (only if we have edited rows)
            if model_name == 'specimen':
                fk_fields_validation, valid_values_cache = build_fk_validation_cache()
                common_errors = {}
                debug_mode = request.POST.get('debug_mode') == 'true'
                
                # Check each edited row for errors (this is fast - only edited rows)
                for idx in edited_rows:
                    row = df.iloc[idx]
                    # Convert row to clean dict for validation
                    row_data = {}
                    for field in expected_fields:
                        if field in fk_fields:
                            col_name = fk_fields[field]
                            if col_name in row:
                                row_data[field] = row[col_name]
                        elif field in row:
                            row_data[field] = row[field]
                    
                    # Create a temporary validation item
                    temp_item = {'errors': [], 'error_fields': set()}
                    validate_specimen_data(row_data, temp_item, common_errors, debug_mode)
                    process_date_fields_unified(row_data, None, None, temp_item, debug_mode)
                    
                    # If this row no longer has errors, count it as fixed
                    if not temp_item['errors']:
                        fixed_errors += 1
                
                # Only do comprehensive scan if we think we might have fixed all errors
                # This avoids the expensive full scan in most cases
                if fixed_errors > 0:
                    # Quick check: if this page still has errors, don't bother with full scan
                    current_page_has_errors = False
                    start_idx_check = (page - 1) * rows_per_page
                    end_idx_check = min(start_idx_check + rows_per_page, len(df))
                    
                    for idx in range(start_idx_check, end_idx_check):
                        if idx in edited_rows:
                            continue  # Skip edited rows, we already checked them
                        
                        row = df.iloc[idx]
                        row_data = {}
                        for field in expected_fields:
                            if field in fk_fields:
                                col_name = fk_fields[field]
                                if col_name in row:
                                    row_data[field] = row[col_name]
                            elif field in row:
                                row_data[field] = row[field]
                        
                        temp_item = {'errors': [], 'error_fields': set()}
                        validate_specimen_data(row_data, temp_item, common_errors, debug_mode)
                        if temp_item['errors']:
                            current_page_has_errors = True
                            break
                    
                    # Only do expensive full scan if current page is clean
                    if not current_page_has_errors:
                        has_any_errors = False
                        # Quick scan through remaining rows (skip current page and edited rows)
                        for idx in range(len(df)):
                            if start_idx_check <= idx < end_idx_check:
                                continue  # Skip current page, already checked
                            
                            row = df.iloc[idx]
                            row_data = {}
                            for field in expected_fields:
                                if field in fk_fields:
                                    col_name = fk_fields[field]
                                    if col_name in row:
                                        row_data[field] = row[col_name]
                                elif field in row:
                                    row_data[field] = row[field]
                            
                            temp_item = {'errors': [], 'error_fields': set()}
                            validate_specimen_data(row_data, temp_item, common_errors, debug_mode)
                            if temp_item['errors']:
                                has_any_errors = True
                                break  # Found error, stop scanning
                        
                        # Update global error state only if comprehensive scan was clean
                        if not has_any_errors:
                            request.session['import_all_errors'] = False
                            request.session['show_errors_only'] = False  # Auto-disable filter
                            request.session.modified = True
        
        # Show intelligent success message
        if edited_count > 0:
            message = f'Applied {edited_count} edit(s) successfully.'
            if fixed_errors > 0:
                message += f' Fixed {fixed_errors} error(s).'
            if not request.session.get('import_all_errors', False):
                message += ' All errors have been resolved!'
            messages.success(request, message)
        else:
            messages.info(request, 'No changes were made.')
        
        # Process the current page with updated data
        return process_paginated_import_data(request, page, rows_per_page, context)
        
    # Step 4: Handle revalidation of form data
    elif request.method == 'POST' and request.POST.get('revalidate'):
        # Get pagination parameters
        page = int(request.POST.get('current_page', 1))
        rows_per_page = int(request.POST.get('rows_per_page', 20))
        
        # Check if we have stored data
        if 'import_df' not in request.session:
            messages.error(request, 'Session expired or invalid data. Please upload the file again.')
            return redirect('import_model', model_name=model_name)
        
        # Apply any current edits first, then revalidate
        df = deserialize_dataframe(request.session.get('import_df'))
        if df is None:
            messages.error(request, 'Session expired or invalid data. Please upload the file again.')
            return redirect('import_model', model_name=model_name)
        
        # Get the row indices for the current page
        total_rows = len(df)
        start_idx = (page - 1) * rows_per_page
        end_idx = min(start_idx + rows_per_page, total_rows)
        
        # Apply any edits from the form to the DataFrame
        fields = request.session.get('import_expected_fields', [])
        fk_fields = request.session.get('import_fk_fields', {})
        
        # Only process rows that actually have form data (important when error filter is active)
        for i, idx in enumerate(range(start_idx, end_idx)):
            # Check if this row has any form data - if not, skip it entirely
            has_form_data = False
            for field in fields:
                form_key = f'row_{i}_{field}'
                if form_key in request.POST:
                    has_form_data = True
                    break
            
            # Skip this row if no form data found (means it wasn't visible in the form)
            if not has_form_data:
                continue
                
            for field in fields:
                form_key = f'row_{i}_{field}'
                # Only process if the form field actually exists in POST data
                if form_key not in request.POST:
                    continue
                    
                new_value = request.POST.get(form_key, '').strip()
                if new_value == '':
                    new_value = None
                    
                # Update the cell in the dataframe
                if field in fk_fields:
                    col_name = fk_fields[field]
                    if col_name in df.columns:
                        df.at[idx, col_name] = new_value
                elif field in df.columns:
                    df.at[idx, field] = new_value
        
        # Save updated dataframe back to session
        request.session['import_df'] = serialize_dataframe(df)
        request.session.modified = True
        
        # Process the current page with updated data
        return process_paginated_import_data(request, page, rows_per_page, context)
        
    # Step 5: Handle import confirmation
    elif request.method == 'POST' and request.POST.get('confirm'):
        # Get stored data from session
        df = deserialize_dataframe(request.session.get('import_df'))
        expected_fields = request.session.get('import_expected_fields', [])
        fk_fields = request.session.get('import_fk_fields', {})
        
        if df is None or not expected_fields:
            messages.error(request, 'Session expired or invalid data. Please upload the file again.')
            return redirect('import_model', model_name=model_name)
        
        # Get model fields for import
        fields = expected_fields
        debug_mode = request.POST.get('debug_mode') == 'true'
        
        # Security check: Ensure there are no validation errors before proceeding
        # This prevents bypassing the disabled button via direct POST
        if request.session.get('import_all_errors', False):
            messages.error(
                request,
                'Import cancelled: There are validation errors that must be fixed before importing.'
            )
            return redirect('import_model', model_name=model_name)
        
        # Initialize counters
        imported_count = 0
        skipped_count = 0
        
        # Clear previous import errors
        if 'import_errors' in request.session:
            del request.session['import_errors']
        request.session['import_errors'] = []
        
        # Process each row from the stored dataframe
        for idx, row in df.iterrows():
            # Convert row to clean dict (None instead of NaN)
            row_data = {}
            for field in fields:
                # For foreign keys, use the mapped column name if available
                if field in fk_fields:
                    col_name = fk_fields[field]
                    if col_name in row:
                        row_data[field] = row[col_name]
                # For regular fields, use the field name directly
                elif field in row:
                    row_data[field] = row[field]
            
            # Clean data: convert empty strings to None
            for field, value in row_data.items():
                if value == '':
                    row_data[field] = None
            
            # Step 5: Process and convert data before creating objects
            try:
                # Note: We don't need to check for duplicates here again, since:
                # 1. The preview validation already caught any duplicates
                # 2. We checked request.session['has_errors'] above and prevented import if errors exist
                # 3. The user had to fix all errors before the import button was enabled
                
                # Model-specific conversions
                if model_name == 'specimen':
                    # Process foreign key fields using our helper function
                    process_foreign_keys(row_data, idx, request, debug_mode)
                    
                    # Process all date fields using our unified helper
                    process_date_fields_unified(row_data, request, idx, None, debug_mode)
                    
                    # Check if event date processing caused a skip flag
                    # The check is already done inside process_date_fields_unified, 
                    # but we need to handle the skip logic here
                    success, should_skip, _ = process_event_date(row_data, None, request, idx, debug_mode)
                    if not success and should_skip:
                        skipped_count += 1
                        raise ValueError("Error processing eventDate, row skipped")
                    
                    # Process eventTime to strip seconds from Excel time values (HH:MM:SS → HH:MM)
                    if 'eventTime' in row_data and row_data['eventTime']:
                        time_value = str(row_data['eventTime'])
                        
                        # Only strip seconds if it's a standard time format (not a range with hyphens or other special formats)
                        if time_value.count(':') == 2 and '-' not in time_value and '/' not in time_value and ',' not in time_value:
                            # Extract just the hours and minutes (HH:MM)
                            try:
                                hour_minute = ':'.join(time_value.split(':')[:2])
                                row_data['eventTime'] = hour_minute
                                if debug_mode and hasattr(request, '_import_context'):
                                    messages.info(request, f"Stripped seconds from time value: '{time_value}' → '{hour_minute}'")
                            except Exception as time_ex:
                                if debug_mode and hasattr(request, '_import_context'):
                                    messages.warning(request, f"Failed to strip seconds from time '{time_value}': {str(time_ex)}")
                    
                    # Handle exact_loc boolean field
                    if 'exact_loc' in row_data and row_data['exact_loc']:
                        # Normalize to TRUE/FALSE strings
                        value = str(row_data['exact_loc']).upper()
                        if value in ['TRUE', 'T', 'YES', 'Y', '1']:
                            row_data['exact_loc'] = 'TRUE'
                        else:
                            row_data['exact_loc'] = 'FALSE'
                    
                    # Convert numeric fields to integers
                    int_fields = ['specimenNumber', 'year', 'month', 'day']
                    for field in int_fields:
                        if field in row_data and row_data[field]:
                            try:
                                # Convert to int, handling potential decimal points
                                if '.' in str(row_data[field]):
                                    row_data[field] = str(int(float(row_data[field])))
                                # Ensure it's a string representation of an integer
                                int(row_data[field])  # Just to validate
                            except (ValueError, TypeError):
                                # If not a valid number, keep as is
                                pass
                
                # Ensure date fields are properly formatted for Django model
                date_fields = ['eventDate', 'dateIdentified', 'georeferencedDate']
                for field in date_fields:
                    if field in row_data and row_data[field]:
                        # If it's a datetime or date object, ensure we get just the date part
                        if hasattr(row_data[field], 'date'):
                            row_data[field] = row_data[field].date()
                        elif hasattr(row_data[field], 'strftime'):
                            # It's already a date object, leave it as is
                            pass
                        elif isinstance(row_data[field], str) and ' ' in row_data[field]:
                            # Handle string with datetime format by extracting just the date part
                            try:
                                # Extract date part from datetime string (before any space)
                                date_part = row_data[field].split(' ')[0]
                                row_data[field] = date_part
                                if debug_mode and hasattr(request, '_import_context'):
                                    messages.info(request, f"Extracted date part '{date_part}' from '{row_data[field]}' for {field}")
                            except Exception as date_ex:
                                if debug_mode and hasattr(request, '_import_context'):
                                    messages.warning(request, f"Failed to extract date part from '{row_data[field]}': {str(date_ex)}")
                
                # Create the object with processed data
                instance = model.objects.create(**row_data)
                
                # Post-creation processing
                if model_name == 'specimen':
                    # Generate catalogNumber if missing
                    if not instance.catalogNumber:
                        year = instance.year or ''
                        locality_code = instance.locality.localityCode if instance.locality else ''
                        specimen_number = instance.specimenNumber or ''
                        
                        # Only generate if we have all components
                        if year and locality_code and specimen_number:
                            instance.catalogNumber = f"{year}-{locality_code}-{specimen_number}"
                            instance.save()
                            
                            if debug_mode and hasattr(request, '_import_context'):
                                messages.info(request, f"Auto-generated catalogNumber '{instance.catalogNumber}' for row {idx+1}")
                
                imported_count += 1
                
            except Exception as e:
                # Handle import error
                error_message = str(e)
                
                # Create a user-friendly error message
                user_message = f"Error in row {idx+1}: "
                
                # Handle common error types
                if "violates not-null constraint" in error_message:
                    field_match = re.search(r'column "([^"]+)"', error_message)
                    field_name = field_match.group(1) if field_match else "unknown field"
                    user_message += f"The field '{field_name}' cannot be empty"
                
                elif "value too long for type" in error_message:
                    field_match = re.search(r'column "([^"]+)"', error_message)
                    field_name = field_match.group(1) if field_match else "unknown field"
                    user_message += f"The value for '{field_name}' is too long"
                    
                elif "duplicate key value violates unique constraint" in error_message:
                    user_message += "This record has a duplicate value for a unique field"
                
                elif "invalid date format" in error_message:
                    field_match = re.search(r'column "([^"]+)"', error_message)
                    field_name = field_match.group(1) if field_match else "date field"
                    user_message += f"Invalid date format for '{field_name}'. Please ensure it's in YYYY-MM-DD format"
                    
                else:
                    # Default to simplified error message
                    user_message += f"Could not import - {error_message}"
                
                # Log the error
                messages.error(request, user_message)
                
                # Store in session for display after redirect
                import_errors = request.session.get('import_errors', [])
                import_errors.append(user_message)
                request.session['import_errors'] = import_errors
                
                # Add technical details in debug mode
                if debug_mode:
                    debug_message = f"Technical details for row {idx+1}: {error_message}"
                    import_errors.append(debug_message)
                
                skipped_count += 1
        
        # Prepare summary message
        if imported_count > 0:
            messages.success(
                request, 
                f'Successfully imported {imported_count} records '
                f'({skipped_count} skipped due to errors).'
            )
        else:
            messages.error(request, 'No records were imported successfully.')
        
        # Handle import errors - make sure they appear on the list page
        import_errors = request.session.get('import_errors', [])
        if import_errors:
            request.session['import_complete'] = True
            # Don't add this warning message as we'll show detailed errors on the list page
            # messages.warning(request, f'There were {len(import_errors)} errors during import. See details below.')
        else:
            # Clear any previous import state
            request.session['import_complete'] = False
            if 'import_errors' in request.session:
                del request.session['import_errors']
            
            # Clean up session data from pagination
            for key in ['import_df', 'import_expected_fields', 'import_fk_fields', 
                        'import_unique_field', 'import_model_name', 'import_page',
                        'import_rows_per_page', 'import_all_errors', 'show_errors_only']:
                if key in request.session:
                    del request.session[key]
        
        # Save the session before redirecting
        request.session.modified = True
        
        # Redirect to the list view
        return redirect('dynamic_list', model_name=model_name)
    
    # Render initial import form
    return render(request, 'butterflies/import_model.html', context)

def process_paginated_import_data(request, page=1, rows_per_page=20, context=None):
    """
    Process a subset of import data for a specific page.
    This function handles the pagination of import preview data.
    
    Parameters:
        request: HTTP request object
        page: Page number to display (1-indexed)
        rows_per_page: Number of rows to display per page
        context: Existing context or None to create a new one
    Returns:
        Rendered template with paginated preview data
    """
    # Ensure import context is set for debug messages
    if not hasattr(request, '_import_context'):
        request._import_context = True
    # Initialize context if not provided
    if context is None:
        context = {}
    
    # Get stored data from session
    df = deserialize_dataframe(request.session.get('import_df'))
    expected_fields = request.session.get('import_expected_fields', [])
    fk_fields = request.session.get('import_fk_fields', {})
    unique_field = request.session.get('import_unique_field')
    model_name = request.session.get('import_model_name')
    
    # Get or create model from name
    model = None
    for m in model_list():
        if m._meta.model_name == model_name:
            model = m
            break
            
    if not model or df is None:
        messages.error(request, 'Session expired or invalid data. Please upload the file again.')
        return redirect('import_model', model_name=model_name)
    
    # Set up pagination
    total_rows = len(df)
    paginator = Paginator(range(total_rows), rows_per_page)
    
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    # Get slice of dataframe for current page
    start_idx = (page_obj.number - 1) * rows_per_page
    end_idx = min(start_idx + rows_per_page, total_rows)
    page_indices = list(range(start_idx, end_idx))
    
    # Get debug mode parameter
    debug_mode = request.POST.get('debug_mode') == 'true'
    print(f"[DEBUG] Debug mode: {debug_mode}")
    
    # Extract rows for this page and convert to dictionaries
    all_rows_data = []
    for idx in page_indices:
        row = df.iloc[idx]
        # Convert row to clean dict (None instead of NaN)
        row_data = {}
        for field in expected_fields:
            # For foreign keys, use the mapped column name if available
            if field in fk_fields:
                col_name = fk_fields[field]
                if col_name in row:
                    row_data[field] = row[col_name]
            # For regular fields, use the field name directly
            elif field in row:
                row_data[field] = row[field]
        
        all_rows_data.append(row_data)
    
    # Pre-validate to identify common error patterns
    common_errors = {}
    if model_name == 'specimen':
        # Get foreign key validation info
        fk_fields_validation, valid_values_cache = build_fk_validation_cache()
        
        # Track invalid values
        for row_data in all_rows_data:
            if 'sex' in row_data and row_data['sex']:
                if row_data['sex'] not in ['male', 'female', '.']:
                    if 'sex_values' not in common_errors:
                        common_errors['sex_values'] = set()
                    common_errors['sex_values'].add(row_data['sex'])
                    
            # Track invalid foreign keys
            for fk_field in fk_fields_validation:
                if fk_field in row_data and row_data[fk_field]:
                    # Special case: allow "." as a placeholder for identifiedBy
                    if fk_field == 'identifiedBy' and row_data[fk_field] == '.':
                        continue
                        
                    if row_data[fk_field] not in valid_values_cache[fk_field]:
                        key = f"{fk_field}_invalid"
                        if key not in common_errors:
                            common_errors[key] = set()
                        common_errors[key].add(row_data[fk_field])
    
    # Now process each row for preview with consistent error detection
    preview_data = []
    for row_index, row_data in enumerate(all_rows_data):
        # Check for duplicates if we have a unique field
        is_duplicate = False
        suggested_key = None
        
        if unique_field and unique_field in row_data and row_data[unique_field]:
            unique_value = row_data[unique_field]
            # Check for duplicates and mark as error
            if model.objects.filter(**{unique_field: unique_value}).exists():
                is_duplicate = True
                # Generate a suggested unique key
                if isinstance(unique_value, str):
                    # Try to find a non-existent suffix
                    suffix = 2
                    while model.objects.filter(**{unique_field: f"{unique_value}-{suffix}"}).exists():
                        suffix += 1
                    suggested_key = f"{unique_value}-{suffix}"
                else:
                    suggested_key = None
        
        # Create preview item
        preview_item = {
            'data': row_data,
            'duplicate': is_duplicate,
            'suggested_key': suggested_key,
            'warnings': [],  # Store any format warnings here
            'errors': [],    # Store validation errors here that must be fixed
            'error_fields': set()  # Track fields with errors for highlighting
        }
        
        # Add duplicate error message if duplicate was found
        if is_duplicate and unique_field:
            unique_value = row_data[unique_field]
            field_label = unique_field.capitalize()
            error_msg = f"Duplicate {field_label}: '{unique_value}' already exists in the database"
            preview_item['errors'].append(error_msg)
            preview_item['error_fields'].add(unique_field)
        
        # Check for potential format issues
        if model_name == 'specimen':
            # Use our unified helper functions for validation
            validate_specimen_data(row_data, preview_item, common_errors, debug_mode)
            
            # Process dates using our unified helper
            process_date_fields_unified(row_data, None, None, preview_item, debug_mode)
        
        # Add the preview item
        preview_data.append(preview_item)
    
    # Add preview data to context
    context['preview'] = preview_data
    context['fields'] = expected_fields
    context['paginator'] = paginator
    context['page_obj'] = page_obj
    context['current_page'] = page_obj.number
    context['has_next'] = page_obj.has_next()
    context['has_previous'] = page_obj.has_previous()
    context['total_pages'] = paginator.num_pages
    context['total_rows'] = total_rows
    
    # Check if any rows have errors that would prevent import (do this first)
    has_errors = False
    error_count = 0
    for item in preview_data:
        if item['errors']:
            has_errors = True
            error_count += len(item['errors'])
    
    # Add error filter state
    show_errors_only = request.session.get('show_errors_only', False)
    
    # Auto-disable error filter if there are no errors on this page or globally
    if show_errors_only and not has_errors and not request.session.get('import_all_errors', False):
        show_errors_only = False
        request.session['show_errors_only'] = False
        request.session.modified = True
        messages.info(request, 'Error filter automatically disabled - no errors found.')
    
    # Update all_errors in session to track if any page has errors
    all_errors = request.session.get('import_all_errors', False)
    
    # If this is a revalidation request, do a comprehensive check across all data
    if request.POST.get('revalidate'):
        # For revalidation, scan the entire dataset to see if any errors remain
        has_any_errors = False
        
        # Quick scan through all rows in the dataframe to check for errors
        for idx in range(len(df)):
            row = df.iloc[idx]
            # Convert row to clean dict for validation
            row_data = {}
            for field in expected_fields:
                if field in fk_fields:
                    col_name = fk_fields[field]
                    if col_name in row:
                        row_data[field] = row[col_name]
                elif field in row:
                    row_data[field] = row[field]
            
            # Quick validation check for this row
            if model_name == 'specimen':
                # Check basic validation that would cause errors
                temp_item = {'errors': [], 'error_fields': set()}
                validate_specimen_data(row_data, temp_item, common_errors, debug_mode)
                if temp_item['errors']:
                    has_any_errors = True
                    break  # Found at least one error, that's enough
        
        # Update the global error state based on comprehensive check
        request.session['import_all_errors'] = has_any_errors
        all_errors = has_any_errors
        
        # Add appropriate message
        if not has_any_errors:
            messages.success(request, 'All errors have been fixed! You can now import the data.')
            # Also disable error filter since there are no more errors
            request.session['show_errors_only'] = False
            show_errors_only = False
        elif has_errors:
            messages.warning(request, f'Still found {error_count} error(s) on this page. Please fix all errors before importing.')
        else:
            messages.info(request, 'This page looks good, but there are still errors on other pages. Please check all pages.')
    else:
        # For regular navigation, maintain the cumulative error state
        request.session['import_all_errors'] = all_errors or has_errors
        all_errors = all_errors or has_errors
    
    # Set the context values after all processing is done
    context['show_errors_only'] = show_errors_only
    context['has_errors'] = has_errors
    context['all_errors'] = all_errors
    context['error_count'] = error_count
    
    # Store pagination info in session
    request.session['import_page'] = page_obj.number
    request.session['import_rows_per_page'] = rows_per_page
    request.session.modified = True
    
    # Add model_name for template
    context['model_name'] = model_name
    context['model_name_internal'] = model_name
    context['model_name_plural'] = model._meta.verbose_name_plural
    context['has_unique_field'] = unique_field is not None
    context['unique_field'] = unique_field
    
    # Render preview template
    return render(request, 'butterflies/import_model_preview.html', context)

# --- List View for All Models ---

@admin_required
def debug_bulk_delete_specimen(request):
    """
    Debug feature to bulk delete specimens.
    Provides a confirmation screen and requires typing "DELETE" to confirm.
    This is a debug-only feature and should be used with caution.
    Requires admin privileges.
    
    Parameters:
        request: HTTP request object
    Returns:
        Confirmation page or redirect to specimen list after deletion
    """
    if request.method == 'POST':
        confirm_text = request.POST.get('confirm_text', '')
        if confirm_text == 'DELETE':
            # Perform the bulk delete
            count = Specimen.objects.all().count()
            Specimen.objects.all().delete()
            messages.success(request, f"Successfully deleted {count} specimen records.")
            return redirect('dynamic_list', model_name='specimen')
        else:
            messages.error(request, "Confirmation text did not match. Deletion canceled.")
            return redirect('dynamic_list', model_name='specimen')
    
    # GET request - show confirmation page
    count = Specimen.objects.all().count()
    context = {
        'count': count,
    }
    return render(request, 'butterflies/bulk_delete_confirm.html', context)

@admin_required
def debug_bulk_delete_specimen_filtered(request):
    """
    Debug feature to bulk delete ONLY the currently filtered specimen records.
    Reuses the same filtering logic as dynamic_list via apply_model_filters.
    Shows a confirmation screen that preserves the current query parameters and
    requires typing "DELETE" to confirm.
    """
    from .filter_utils import apply_model_filters

    # Build filtered queryset exactly like dynamic_list does
    model = Specimen
    queryset = model.objects.all()

    special_filters = {
        'catalogNumber': {'field': 'catalogNumber', 'range_support': True},
        'locality': {'field': 'locality__localityCode', 'range_support': False},
        'specimenNumber': {'field': 'specimenNumber', 'range_support': True},
        'year': {'field': 'year', 'range_support': True},
    }

    filtered_qs = apply_model_filters(queryset, model, request, special_filters)

    if request.method == 'POST':
        confirm_text = request.POST.get('confirm_text', '')
        if confirm_text == 'DELETE':
            count = filtered_qs.count()
            filtered_qs.delete()
            messages.success(request, f"Successfully deleted {count} filtered specimen records.")
            return redirect('dynamic_list', model_name='specimen')
        else:
            messages.error(request, "Confirmation text did not match. Deletion canceled.")
            return redirect('dynamic_list', model_name='specimen')

    # GET request - show confirmation page with count and preserve filters
    count = filtered_qs.count()
    # Keep the original query string to round-trip on the form action
    query_string = request.META.get('QUERY_STRING', '')
    context = {
        'count': count,
        'query_string': query_string,
        'filtered': True,
    }
    return render(request, 'butterflies/bulk_delete_confirm.html', context)

def all_list(request):
    """
    Display objects from all models in the butterflies app.
    Supports multi-field search across all models.
    
    Parameters:
        request: HTTP request object
    Returns:
        Rendered template with tables for all models
    """
    app_models = apps.get_app_config('butterflies').get_models()
    model_tables = []
    for model in app_models:
        # Get all objects for the model
        objects = model.objects.all()
        # Get all fields except id
        fields = [
            {'name': field.name, 'verbose_name': getattr(field, 'verbose_name', field.name)}
            for field in model._meta.fields
            if field.name != 'id'
        ]
        # Multi-field search for each model
        for field in fields:
            value = request.GET.get(field['name'])
            if value:
                objects = objects.filter(**{f"{field['name']}__icontains": value})
        # Attach model_name to each object for template use
        obj_list = []
        for obj in objects:
            obj.model_name_internal = model._meta.model_name
            obj_list.append(obj)
        model_tables.append({
            'model_name': model._meta.verbose_name.title(),
            'model_name_internal': model._meta.model_name,
            'objects': obj_list,
            'fields': fields,
        })
    return render(request, 'butterflies/all_list.html', {
        'model_tables': model_tables,
        'request': request,
    })

# --- Session Serialization Helpers ---
def serialize_dataframe(df):
    """
    Serialize a pandas DataFrame to store in session.
    
    Parameters:
        df: pandas DataFrame to serialize
    Returns:
        str: Base64 encoded serialized DataFrame
    """
    pickled = pickle.dumps(df)
    return base64.b64encode(pickled).decode('utf-8')

def deserialize_dataframe(serialized_df):
    """
    Deserialize a pandas DataFrame from session storage.
    
    Parameters:
        serialized_df: Base64 encoded serialized DataFrame
    Returns:
        pandas DataFrame
    """
    if not serialized_df:
        return None
    
    try:
        return pickle.loads(base64.b64decode(serialized_df.encode('utf-8')))
    except Exception as e:
        print(f"Error deserializing DataFrame: {str(e)}")
        return None

def custom_logout(request):
    """
    Custom logout view that supports both GET and POST requests.
    Logs out the user and redirects to the logged_out template.
    Also clears guest mode if active.
    
    Parameters:
        request: HTTP request object
    Returns:
        Rendered logged_out template
    """
    # Clear guest mode if active
    if 'guest_mode' in request.session:
        del request.session['guest_mode']
    logout(request)
    return render(request, 'butterflies/auth/logged_out.html')

def guest_login(request):
    """
    View to enable guest mode for read-only access.
    Sets a session flag to allow access without authentication.
    
    Parameters:
        request: HTTP request object
    Returns:
        Redirect to next page or home page
    """
    # Set guest mode in session
    request.session['guest_mode'] = True
    
    # Always redirect to guest_view (explicit requirement)
    from django.urls import reverse
    next_url = reverse('guest_view')
    
    # Add a message to inform user about guest limitations
    messages.info(request, 
        'You are now browsing as a guest. You can view all data but cannot make any changes. '
        'To edit data, please log in with your credentials.')
    
    return redirect(next_url)

def guest_logout(request):
    """
    View to disable guest mode and return to login page.
    
    Parameters:
        request: HTTP request object
    Returns:
        Redirect to login page
    """
    # Clear guest mode
    if 'guest_mode' in request.session:
        del request.session['guest_mode']
    
    messages.info(request, 'Guest session ended.')
    return redirect('/accounts/login/')